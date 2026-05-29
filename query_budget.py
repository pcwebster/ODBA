# =============================================================================
# ODBA -- Open Defense Budget Analytics
# query_budget.py  |  FY2027 Budget Report Generator
# =============================================================================
# Reads fact_budget_line_items.parquet and writes a multi-tab Excel workbook:
#   1. Summary          -- total funding by appropriation type
#   2. By Agency        -- FY2027 totals ranked by agency
#   3. Year-over-Year   -- prior year vs FY2027, dollar + % change
#   4. RDT&E by BA      -- R&D budget activity breakdown (BA1-BA8)
#   5. Procurement      -- P-1 line items with full funding profile
#   6. RDT&E            -- R-1 program elements with full funding profile
#   7. All Records      -- complete data dump (all 33 columns)
#
# Output: output/ODBA_FY2027_Budget_Report.xlsx
# =============================================================================

import sys
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent.resolve()
PARQUET_FILE = SCRIPT_DIR / "output" / "fact_budget_line_items.parquet"
OUTPUT_DIR   = SCRIPT_DIR / "output"
OUTPUT_FILE  = OUTPUT_DIR / "ODBA_FY2027_Budget_Report.xlsx"

# ── Colors (DoD blue palette) ─────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")   # medium blue
ALT_ROW_FILL  = PatternFill("solid", fgColor="DCE6F1")   # light blue
TOTAL_FILL    = PatternFill("solid", fgColor="BDD7EE")   # accent blue
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
NEG_FILL      = PatternFill("solid", fgColor="FCE4D6")   # salmon for decreases

HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT  = Font(bold=True, color="FFFFFF", size=10)
TOTAL_FONT    = Font(bold=True, size=10)
BODY_FONT     = Font(size=10)
TITLE_FONT    = Font(bold=True, size=14, color="1F3864")

THIN = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMT  = '#,##0.0'      # millions with one decimal
PCT_FMT  = '+0.0%;-0.0%;0.0%'
DELTA_FMT = '+#,##0.0;-#,##0.0;0.0'


# =============================================================================
# Helpers
# =============================================================================

def autofit(ws, min_width=8, max_width=60):
    """Set column widths based on content."""
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def write_df(ws, df, start_row=1, header_fill=None, alt_rows=True,
             number_cols=None, pct_cols=None, delta_cols=None):
    """
    Write a DataFrame to a worksheet with formatting.
    Returns the row number after the last data row.
    """
    hfill = header_fill or HEADER_FILL
    number_cols = number_cols or []
    pct_cols    = pct_cols    or []
    delta_cols  = delta_cols  or []

    # Header row
    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.font   = HEADER_FONT
        cell.fill   = hfill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for r_idx, row in enumerate(df.itertuples(index=False), 1):
        fill = ALT_ROW_FILL if (alt_rows and r_idx % 2 == 0) else WHITE_FILL
        for c_idx, (col_name, value) in enumerate(zip(df.columns, row), 1):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=value)
            cell.font   = BODY_FONT
            cell.fill   = fill
            cell.border = BORDER
            if col_name in number_cols:
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")
            elif col_name in pct_cols:
                cell.number_format = PCT_FMT
                cell.alignment = Alignment(horizontal="right")
                if isinstance(value, (int, float)) and value < 0:
                    cell.fill = NEG_FILL
            elif col_name in delta_cols:
                cell.number_format = DELTA_FMT
                cell.alignment = Alignment(horizontal="right")
                if isinstance(value, (int, float)) and value < 0:
                    cell.fill = NEG_FILL
            else:
                cell.alignment = Alignment(wrap_text=False)

    return start_row + len(df) + 1


def add_title(ws, title, subtitle=None, row=1):
    """Write a sheet title + optional subtitle."""
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    if subtitle:
        ws.cell(row=row + 1, column=1, value=subtitle).font = Font(italic=True, size=10, color="595959")
    return row + (3 if subtitle else 2)


def freeze_top(ws, data_start_row):
    """Freeze the header row."""
    ws.freeze_panes = ws.cell(row=data_start_row + 1, column=1)


# =============================================================================
# Queries
# =============================================================================

def run_queries(parquet_path):
    con = duckdb.connect()
    p = str(parquet_path).replace("\\", "/")
    # DBDP-50: bind path as parameter instead of f-string interpolation
    # DuckDB supports ? placeholders: con.execute("... read_parquet(?) ...", [p])

    print("  Running queries...")

    # 1. Summary by appropriation type
    q_summary = con.execute("""
        SELECT
            appropriation_type                         AS "Appropriation Type",
            exhibit_type                               AS "Exhibit",
            COUNT(*)                                   AS "# Line Items",
            ROUND(SUM(cost_prior_year),  1)            AS "FY2026 Enacted ($M)",
            ROUND(SUM(cost_fy2027),      1)            AS "FY2027 Request ($M)",
            ROUND(SUM(cost_fy2027) - SUM(cost_prior_year), 1) AS "Change ($M)",
            CASE WHEN SUM(cost_prior_year) > 0
                 THEN (SUM(cost_fy2027) - SUM(cost_prior_year)) / SUM(cost_prior_year)
                 ELSE NULL END                         AS "Change (%)"
        FROM read_parquet(?)
        WHERE cost_prior_year IS NOT NULL OR cost_fy2027 IS NOT NULL
        GROUP BY 1, 2
        ORDER BY "FY2027 Request ($M)" DESC NULLS LAST
    """, [p]).df()

    # 2. By agency
    q_agency = con.execute("""
        SELECT
            service_agency_acronym                     AS "Agency",
            service_agency_name                        AS "Agency Name",
            appropriation_type                         AS "Appropriation",
            COUNT(*)                                   AS "# Line Items",
            ROUND(SUM(cost_prior_year), 1)             AS "FY2026 Enacted ($M)",
            ROUND(SUM(cost_fy2027),     1)             AS "FY2027 Request ($M)",
            ROUND(SUM(cost_fy2027) - SUM(cost_prior_year), 1) AS "Change ($M)",
            CASE WHEN SUM(cost_prior_year) > 0
                 THEN (SUM(cost_fy2027) - SUM(cost_prior_year)) / SUM(cost_prior_year)
                 ELSE NULL END                         AS "Change (%)"
        FROM read_parquet(?)
        GROUP BY 1, 2, 3
        ORDER BY "FY2027 Request ($M)" DESC NULLS LAST
    """, [p]).df()

    # 3. Year-over-year by line item (XML records with both years populated)
    q_yoy = con.execute("""
        SELECT
            service_agency_acronym                     AS "Agency",
            appropriation_type                         AS "Type",
            COALESCE(NULLIF(line_item_title,''), line_item_title) AS "Line Item / Program",
            program_element                            AS "PE / Line #",
            budget_activity_number                     AS "BA",
            ROUND(cost_prior_year, 1)                  AS "FY2026 Enacted ($M)",
            ROUND(cost_fy2027,     1)                  AS "FY2027 Request ($M)",
            ROUND(cost_fy2027 - cost_prior_year, 1)    AS "Change ($M)",
            CASE WHEN cost_prior_year > 0
                 THEN (cost_fy2027 - cost_prior_year) / cost_prior_year
                 ELSE NULL END                         AS "Change (%)"
        FROM read_parquet(?)
        WHERE file_format = 'XML'
          AND cost_prior_year IS NOT NULL
          AND cost_fy2027     IS NOT NULL
        ORDER BY ABS(cost_fy2027 - cost_prior_year) DESC NULLS LAST
    """, [p]).df()

    # 4. RDT&E by Budget Activity
    q_ba = con.execute("""
        SELECT
            budget_activity_number                     AS "BA",
            budget_activity_title                      AS "Budget Activity Title",
            COUNT(*)                                   AS "# Programs",
            ROUND(SUM(cost_prior_year), 1)             AS "FY2026 Enacted ($M)",
            ROUND(SUM(cost_fy2027),     1)             AS "FY2027 Request ($M)",
            ROUND(SUM(cost_fy2028),     1)             AS "FY2028 FYDP ($M)",
            ROUND(SUM(cost_fy2029),     1)             AS "FY2029 FYDP ($M)",
            ROUND(SUM(cost_fy2030),     1)             AS "FY2030 FYDP ($M)",
            ROUND(SUM(cost_fy2031),     1)             AS "FY2031 FYDP ($M)"
        FROM read_parquet(?)
        WHERE appropriation_type = 'RDT&E'
        GROUP BY 1, 2
        ORDER BY CAST(budget_activity_number AS INTEGER) NULLS LAST
    """, [p]).df()

    # 5. Procurement detail
    q_proc = con.execute("""
        SELECT
            service_agency_acronym                     AS "Agency",
            line_item_number                           AS "Line #",
            line_item_title                            AS "Line Item Title",
            budget_activity_number                     AS "BA",
            budget_activity_title                      AS "BA Title",
            ROUND(cost_all_prior_years, 1)             AS "All Prior ($M)",
            ROUND(cost_prior_year,  1)                 AS "FY2026 ($M)",
            ROUND(cost_current_year,1)                 AS "FY2026 Enacted ($M)",
            ROUND(cost_fy2027,      1)                 AS "FY2027 ($M)",
            ROUND(cost_fy2028,      1)                 AS "FY2028 ($M)",
            ROUND(cost_fy2029,      1)                 AS "FY2029 ($M)",
            ROUND(cost_fy2030,      1)                 AS "FY2030 ($M)",
            ROUND(cost_fy2031,      1)                 AS "FY2031 ($M)",
            source_file                                AS "Source File"
        FROM read_parquet(?)
        WHERE appropriation_type = 'Procurement'
        ORDER BY service_agency_acronym, budget_activity_number, line_item_number
    """, [p]).df()

    # 6. RDT&E detail
    q_rdte = con.execute("""
        SELECT
            service_agency_acronym                     AS "Agency",
            program_element                            AS "PE Number",
            line_item_title                            AS "Program Element Title",
            budget_activity_number                     AS "BA",
            budget_activity_title                      AS "BA Title",
            ROUND(cost_prior_year,  1)                 AS "FY2026 ($M)",
            ROUND(cost_fy2027,      1)                 AS "FY2027 ($M)",
            ROUND(cost_fy2028,      1)                 AS "FY2028 ($M)",
            ROUND(cost_fy2029,      1)                 AS "FY2029 ($M)",
            ROUND(cost_fy2030,      1)                 AS "FY2030 ($M)",
            ROUND(cost_fy2031,      1)                 AS "FY2031 ($M)",
            source_file                                AS "Source File"
        FROM read_parquet(?)
        WHERE appropriation_type = 'RDT&E'
        ORDER BY service_agency_acronym, budget_activity_number, program_element
    """, [p]).df()

    # 7. All records
    q_all = con.execute("""
        SELECT
            record_id, budget_year, submission_date,
            service_agency_acronym AS "agency_acronym",
            service_agency_name    AS "agency_name",
            appropriation_type, exhibit_type,
            line_item_number, line_item_title,
            program_element,
            budget_activity_number, budget_activity_title,
            ROUND(cost_prior_year,  1) AS cost_prior_year,
            ROUND(cost_fy2027,      1) AS cost_fy2027,
            ROUND(cost_fy2028,      1) AS cost_fy2028,
            ROUND(cost_fy2029,      1) AS cost_fy2029,
            ROUND(cost_fy2030,      1) AS cost_fy2030,
            ROUND(cost_fy2031,      1) AS cost_fy2031,
            cost_units, file_format, source_file
        FROM read_parquet(?)
        ORDER BY appropriation_type, service_agency_acronym, line_item_number
    """, [p]).df()

    return {
        "Summary":        q_summary,
        "By Agency":      q_agency,
        "Year-over-Year": q_yoy,
        "RDT&E by BA":    q_ba,
        "Procurement":    q_proc,
        "RDT&E":          q_rdte,
        "All Records":    q_all,
    }


# =============================================================================
# Build workbook
# =============================================================================

def build_workbook(sheets):
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    num_money = ["FY2026 Enacted ($M)", "FY2027 Request ($M)",
                 "Change ($M)", "FY2026 ($M)", "FY2027 ($M)",
                 "FY2028 ($M)", "FY2028 FYDP ($M)",
                 "FY2029 ($M)", "FY2029 FYDP ($M)",
                 "FY2030 ($M)", "FY2030 FYDP ($M)",
                 "FY2031 ($M)", "FY2031 FYDP ($M)",
                 "All Prior ($M)", "FY2026 Enacted ($M)",
                 "cost_prior_year", "cost_fy2027", "cost_fy2028",
                 "cost_fy2029", "cost_fy2030", "cost_fy2031",
                 "# Programs", "# Line Items",
                 "FY2026 Enacted ($M)"]
    pct_money  = ["Change (%)"]
    delta_money = ["Change ($M)"]

    sheet_meta = {
        "Summary":        ("FY2027 Defense Budget -- Summary",
                           "Total funding by appropriation type  |  Dollars in Millions"),
        "By Agency":      ("FY2027 Defense Budget -- By Agency",
                           "FY2027 request ranked by agency  |  Dollars in Millions"),
        "Year-over-Year": ("Year-over-Year Change by Line Item",
                           "XML records with both FY2026 enacted and FY2027 request  |  Sorted by absolute change"),
        "RDT&E by BA":    ("RDT&E Budget Activity Summary (BA 1-8)",
                           "Includes FYDP profile through FY2031  |  Dollars in Millions"),
        "Procurement":    ("Procurement (P-1) Line Items",
                           "All P-1 procurement line items with FYDP  |  Dollars in Millions"),
        "RDT&E":          ("RDT&E (R-1) Program Elements",
                           "All R-1 program elements with FYDP  |  Dollars in Millions"),
        "All Records":    ("Complete Data -- All Records",
                           "All 33 schema columns  |  Use filters to explore"),
    }

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        title, subtitle = sheet_meta[sheet_name]
        data_start = add_title(ws, title, subtitle)
        write_df(ws, df, start_row=data_start,
                 number_cols=num_money,
                 pct_cols=pct_money,
                 delta_cols=delta_money)
        freeze_top(ws, data_start)
        autofit(ws)

        # Add auto-filter
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A{data_start}:{last_col}{data_start}"

    return wb


# =============================================================================
# Main
# =============================================================================

def main():
    print()
    print("=" * 65)
    print("  ODBA -- FY2027 Budget Report Generator")
    print("=" * 65)
    print()

    if not PARQUET_FILE.exists():
        print(f"  [ERROR] Parquet file not found:")
        print(f"    {PARQUET_FILE}")
        print()
        print("  Run Run-ETL.ps1 first to generate the data file.")
        sys.exit(1)

    OUTPUT_DIR.mkdir(exist_ok=True)

    print(f"  Source : {PARQUET_FILE.name}")
    print(f"  Output : {OUTPUT_FILE.name}")
    print()

    sheets = run_queries(PARQUET_FILE)

    total_rows = sum(len(df) for df in sheets.values())
    print(f"  Building Excel workbook ({len(sheets)} sheets, {total_rows:,} total rows)...")

    wb = build_workbook(sheets)
    wb.save(OUTPUT_FILE)

    size_kb = round(OUTPUT_FILE.stat().st_size / 1024, 1)
    print()
    print("=" * 65)
    print("  Done!")
    print(f"  Report saved: {OUTPUT_FILE}")
    print(f"  Size: {size_kb} KB  |  Sheets: {len(sheets)}")
    print()
    print("  Tabs in the workbook:")
    for name, df in sheets.items():
        print(f"    {name:<20} {len(df):>4} rows")
    print("=" * 65)
    print()


if __name__ == "__main__":
    main()
